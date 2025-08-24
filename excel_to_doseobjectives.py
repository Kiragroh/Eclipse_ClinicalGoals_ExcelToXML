#!/usr/bin/env python3
"""
Simplified Excel-to-XML converter for Clinical Goals templates.

- Input: An Excel (.xlsx) with a sheet named 'Constraints' and the columns listed below.
- Output: TPS-compatible DoseObjectives XML.
- Priority is taken directly from the Excel (no additional rules applied).

Usage:
  python excel_to_doseobjectives.py <input.xlsx> <output.xml> [PreviewID]

If PreviewID is omitted, the Excel file stem is used.
"""
from __future__ import annotations

import sys
import re
from pathlib import Path
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, ElementTree

try:
    from openpyxl import load_workbook
except Exception as e:
    print("openpyxl is required for this simplified converter. Please install it: pip install openpyxl")
    raise

# Configuration
# Eclipse users who should see/use the Clinical Goals in TPS (comma-separated for multiple)
# Example: "domain\\user1,domain\\user2". Leave empty and set in Eclipse after import if unsure.
ASSIGNED_USERS = ""

# Expected column headers
COL_STRUCTURE_IDS = "Structure IDs"
COL_STRUCTURE_CODES = "Structure Codes"
COL_ID_ALIASES = "IDAliases"
COL_DVH_OBJ = "DVH Objective"
COL_EVAL_POINT = "Evaluation Point"
COL_VARIATION = "Variation"
COL_PRIORITY = "Priority"
# Informational columns (not exported to XML fields)
COL_SOURCE = "Source"
COL_TEMPLATE_ID = "TemplateID"
COL_ZUSATZINFO = "ZusatzInfo"  # accept lowercase variant when reading
COL_ENDPOINT_GEQ3 = "Endpoint (grade >= 3)"

# Coding system constants (optional, used if Structure Codes contain a numeric code)
CODE_SCHEME = "FMA"
CODE_SCHEME_VERSION = "3.2"

# Alias selection mode for MeasureItem IDs
#   - "structure_ids": use tokens from Structure IDs (pipe '|' separated)
#   - "idalias_first": use only the first token from IDAliases
#   - "idalias_all": use all tokens from IDAliases (pipe '|' separated)
#   - "all": combine Structure IDs tokens and all IDAliases tokens
# Default: "structure_ids"
ALIAS_MODE = "structure_ids"
# If true, also add the template/preview ID (usually the filename stem) as a MeasureItem alias
# Default False to avoid duplicate-looking items in XML
ADD_PREVIEW_ID_ALIAS = False


def round1(x):
    try:
        return float(f"{float(x):.1f}")
    except Exception:
        return None


def ceil1(x):
    try:
        import math
        xv = float(x)
        return math.ceil(xv * 10.0) / 10.0
    except Exception:
        return None


def first_numeric_code(s: str | None):
    if not s:
        return None
    for part in str(s).split("|"):
        part = part.strip()
        if part and part.isdigit() and int(part) > 0:
            return part
    return None


# Metric parsing (kept consistent with repo's main script)
D_METRIC_TYPES = {
    'Dmean': 8,
    'Dmax': 6,
    'Dmin': 7,
}


def parse_metric(dvh_str: str | None):
    if not dvh_str:
        return None
    s = str(dvh_str).strip()

    if re.search(r"\b(CI|HI|GI|CV)\b", s, re.IGNORECASE):
        return None

    m = re.match(r"^(Mean|Max|Min)\s*\[(Gy|%)\]$", s, re.IGNORECASE)
    if m:
        kind_word = m.group(1).lower()
        unit = m.group(2)
        absolute = (unit.lower() == 'gy')
        type_spec = 0 if absolute else 100
        name = 'Dmean'
        if kind_word == 'max':
            name = 'Dmax'
        elif kind_word == 'min':
            name = 'Dmin'
        tcode = D_METRIC_TYPES.get(name, D_METRIC_TYPES['Dmean'])
        return {"kind": "D", "name": name, "type_code": tcode, "type_specifier": type_spec, "absolute_units": absolute}

    m = re.match(r"^V\s*([0-9]+(?:\.[0-9]+)?)\s*Gy\s*\[(%|cc)\]$", s, re.IGNORECASE)
    if m:
        dose_gy = round1(m.group(1))
        unit = m.group(2)
        if unit == '%':
            return {"kind": "V", "name": "V", "type_code": 3, "type_specifier": dose_gy, "absolute_units": False, "value_transform": None}
        else:
            return {"kind": "V", "name": "V", "type_code": 3, "type_specifier": dose_gy, "absolute_units": True, "value_transform": "cc_to_mm3"}

    m = re.match(r"^D\s*([0-9]+(?:\.[0-9]+)?)\s*(cc|%)\s*\[(Gy|%)\]$", s, re.IGNORECASE)
    if m:
        raw_vol = m.group(1)
        vunit = m.group(2)
        out_unit = m.group(3)
        vol = ceil1(raw_vol) if vunit == 'cc' else round1(raw_vol)
        if vunit == 'cc':
            if out_unit.lower() == 'gy':
                return {"kind": "DV", "name": "D_at_cc", "type_code": 5, "type_specifier": vol, "absolute_units": True}
            if out_unit == '%':
                return {"kind": "DV", "name": "D_at_cc", "type_code": 5, "type_specifier": vol, "absolute_units": False}
        if vunit == '%':
            if out_unit == '%':
                return {"kind": "DV", "name": "D_at_percent", "type_code": 4, "type_specifier": vol, "absolute_units": False}
            if out_unit.lower() == 'gy':
                return {"kind": "DV", "name": "D_at_percent", "type_code": 4, "type_specifier": vol, "absolute_units": True}

    return None


def parse_eval_point(s: str | None):
    """Return (modifier:int, value:float) or (None, None) if invalid.
    Supported patterns (examples):
      - "<=30" -> modifier 3 (<=), value 30
      - ">=95" -> modifier 4 (>=), value 95
      - "=95"  -> modifier 2 (=), value 95
    """
    if not s:
        return None, None
    s = str(s).strip()
    if m := re.match(r"^<=\s*([0-9]+(?:\.[0-9]+)?)$", s):
        return 3, round1(m.group(1))
    if m := re.match(r"^>=\s*([0-9]+(?:\.[0-9]+)?)$", s):
        return 4, round1(m.group(1))
    if m := re.match(r"^=\s*([0-9]+(?:\.[0-9]+)?)$", s):
        return 2, round1(m.group(1))
    return None, None


def load_excel_rows(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Constraints"]
    # Build header map
    headers = {}
    for j, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
        if cell is None:
            continue
        headers[str(cell).strip()] = j
    def val(row_idx: int, col_name: str, default: str = ""):
        idx = headers.get(col_name)
        if idx is None and col_name == COL_ZUSATZINFO:
            idx = headers.get("Zusatzinfo")  # accept lowercase variant
        if idx is None:
            return default
        v = ws.cell(row=row_idx, column=idx).value
        return "" if v is None else v

    rows = []
    for i in range(2, ws.max_row + 1):
        rid = {
            COL_STRUCTURE_IDS: val(i, COL_STRUCTURE_IDS),
            COL_STRUCTURE_CODES: val(i, COL_STRUCTURE_CODES),
            COL_ID_ALIASES: val(i, COL_ID_ALIASES),
            COL_DVH_OBJ: val(i, COL_DVH_OBJ),
            COL_EVAL_POINT: val(i, COL_EVAL_POINT),
            COL_VARIATION: val(i, COL_VARIATION),
            COL_PRIORITY: val(i, COL_PRIORITY),
            COL_SOURCE: val(i, COL_SOURCE),
            COL_TEMPLATE_ID: val(i, COL_TEMPLATE_ID),
            COL_ZUSATZINFO: val(i, COL_ZUSATZINFO),
            COL_ENDPOINT_GEQ3: val(i, COL_ENDPOINT_GEQ3),
        }
        # skip empty DVH Objective rows
        if not str(rid[COL_DVH_OBJ]).strip():
            continue
        rows.append(rid)
    return rows


def build_xml_from_excel(input_xlsx: Path, output_xml: Path, preview_id: str | None = None):
    preview_id = preview_id or input_xlsx.stem

    # Root and Preview
    root = Element("DoseObjectives", {
        "Version": "1.0",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
    })
    now = datetime.now()
    last_modified = now.strftime(" %B %d %Y %H:%M:%S:%f")[:-3]
    description = f"Source Excel: {input_xlsx.name} | Script: excel_to_doseobjectives.py | Converted: {now:%Y-%m-%d %H:%M:%S}"
    SubElement(root, "Preview", {
        "Version": "1.2",
        "ID": preview_id,
        "Type": "DoseObjectives",
        "ApprovalStatus": "Unapproved",
        "Diagnosis": "",
        "TreatmentSite": "",
        "Description": description,
        "AssignedUsers": ASSIGNED_USERS,
        "LastModified": last_modified,
        "ApprovalHistory": f"Created [ {last_modified} ]",
    })

    prescription = SubElement(root, "Prescription", {"Version": "1.10"})

    for row in load_excel_rows(input_xlsx):
        metric = parse_metric(str(row.get(COL_DVH_OBJ) or '').strip())
        if not metric:
            continue
        modifier, value = parse_eval_point(str(row.get(COL_EVAL_POINT) or '').strip())
        if modifier is None or value is None:
            continue

        aliases_raw = str(row.get(COL_ID_ALIASES) or '').strip()
        sids_str = str(row.get(COL_STRUCTURE_IDS) or '').strip()
        idalias_tokens = [a.strip() for a in aliases_raw.split('|') if a.strip()] if aliases_raw else []
        sid_tokens = [a.strip() for a in sids_str.split('|') if a.strip()] if sids_str else []

        if ALIAS_MODE == "idalias_first":
            alias_tokens = idalias_tokens[:1] if idalias_tokens else (sid_tokens[:1] if sid_tokens else [])
        elif ALIAS_MODE == "idalias_all":
            alias_tokens = idalias_tokens if idalias_tokens else (sid_tokens if sid_tokens else [])
        elif ALIAS_MODE == "all":
            alias_tokens = list(dict.fromkeys(sid_tokens + idalias_tokens))
        else:  # "structure_ids" (default)
            alias_tokens = sid_tokens if sid_tokens else (idalias_tokens[:1] if idalias_tokens else [])

        # Require at least one alias; if empty, skip row
        if not alias_tokens:
            continue

        # Optionally append preview_id as extra alias
        if ADD_PREVIEW_ID_ALIAS:
            alias_tokens.append(preview_id)
        seen = set()
        alias_tokens = [a for a in alias_tokens if not (a in seen or seen.add(a))]

        # Gather once-per-row values
        scode = first_numeric_code(str(row.get(COL_STRUCTURE_CODES) or '').strip())
        variation_raw = str(row.get(COL_VARIATION) or '').strip()
        priority_raw = row.get(COL_PRIORITY)
        try:
            priority = int(float(priority_raw)) if str(priority_raw).strip() else 4
        except Exception:
            priority = 4

        for item_id in alias_tokens:
            mi = SubElement(prescription, "MeasureItem", {"ID": item_id})

            if scode:
                sc = SubElement(mi, "StructureCode")
                sc.set("Code", scode)
                sc.set("CodeScheme", CODE_SCHEME)
                sc.set("CodeSchemeVersion", CODE_SCHEME_VERSION)

            SubElement(mi, "Type").text = str(metric["type_code"])  # int -> string
            SubElement(mi, "Modifier").text = str(modifier)

            vtext = value
            if metric.get("value_transform") == "cc_to_mm3" and vtext is not None:
                try:
                    vtext = round1(vtext * 1000.0)
                    if hasattr(vtext, 'is_integer') and vtext.is_integer():
                        vtext = int(vtext)
                except Exception:
                    pass
            SubElement(mi, "Value").text = str(vtext)

            if metric.get("type_specifier") is not None:
                SubElement(mi, "TypeSpecifier").text = str(metric["type_specifier"])  # Gy level

            SubElement(mi, "ReportDQPValueInAbsoluteUnits").text = "true" if metric["absolute_units"] else "false"

            SubElement(mi, "Priority").text = str(priority)

            if variation_raw:
                try:
                    var_val = round1(variation_raw)
                    if metric.get("value_transform") == "cc_to_mm3" and var_val is not None:
                        var_val = round1(var_val * 1000.0)
                        try:
                            if hasattr(var_val, 'is_integer') and var_val.is_integer():
                                var_val = int(var_val)
                        except Exception:
                            pass
                    if var_val is not None:
                        SubElement(mi, "VariationAcceptable").text = str(var_val)
                except Exception:
                    pass

            SubElement(mi, "PrimaryClinicalGoal").text = "false"

    tree = ElementTree(root)
    output_xml.parent.mkdir(parents=True, exist_ok=True)
    tree.write(output_xml, encoding="utf-8", xml_declaration=True)


if __name__ == "__main__":
    # Batch mode fallback: if no args, process all .xlsx in Release/templates/
    if len(sys.argv) < 3:
        script_dir = Path(__file__).resolve().parent
        templates_dir = script_dir / "templates"
        if not templates_dir.exists():
            print("Usage: python excel_to_doseobjectives.py <input.xlsx> <output.xml> [PreviewID]\n"
                  "Or run without arguments from the Release folder to batch convert all Excel files in Release/templates/.")
            sys.exit(1)
        count = 0
        for xlsx in sorted(templates_dir.glob("*.xlsx")):
            out_xml = xlsx.with_suffix('.xml')
            build_xml_from_excel(xlsx, out_xml, xlsx.stem)
            print(f"Written: {out_xml}")
            count += 1
        if count == 0:
            print(f"No .xlsx files found in {templates_dir}")
        sys.exit(0)

    input_xlsx = Path(sys.argv[1])
    output_xml = Path(sys.argv[2])
    preview_id = sys.argv[3] if len(sys.argv) > 3 else None
    build_xml_from_excel(input_xlsx, output_xml, preview_id)
    print(f"Written: {output_xml}")
